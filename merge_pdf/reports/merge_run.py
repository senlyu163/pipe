import PyPDF2
from typing import List, Union
from openpyxl import load_workbook
import os
import tqdm

class EXCEL(object):
    def __init__(self, filename: str, read_only: bool = True) -> None:
        """Initiate EXCEL class
        
        Keyword arguments:
        filename -- The path of excel file.
        read_only -- Read/Write mode. Default read mode to protect raw data.
        Return: None
        """
        self._xl = load_workbook(filename, read_only)
        self._curr_ws = None

    def _check_ws(self):
        if self._curr_ws is None:
            raise RuntimeError('Please active worksheet first')

    @property
    def get_sheets_name(self) -> List[str]:
        return self._xl.sheetnames

    def active_ws_by_name(self, ws_name: str) -> None:
        self._curr_ws = self._xl[ws_name]

    def get_element(self, row: str, col: int):
        return self._curr_ws[row + str(col)].value

    def elements_slice(self, start: str, end: str) -> List[Union[int, float, str]]:
        """Get elements of cell, action like built-in range function in python.
        Notion: [start:end] means double closed region.
        
        Keyword arguments:
        start -- Beginning position of the slice.
        end -- End position of the slice.
        Return: List which contain int/float/str elements in the cells.
        """
        elements = []
        ws_range = "{}:{}".format(start, end)
        for row in self._curr_ws[ws_range]:
            for cell in row:
                elements.append(cell.value)
        return elements
    
    @property
    def rows(self) -> int:
        self._check_ws()
        return len(self._curr_ws['A'])


def get_pos_name_pair(excel: EXCEL, name: str) -> str:
    begin_with = 2
    name_lst = excel.elements_slice("D{}".format(begin_with), "D{}".format(excel.rows))
    name_index = name_lst.index(name)
    pos = excel.get_element("C", name_index + begin_with)
    return "{}-{}".format(pos, name)

def main():
    CWD = os.getcwd()
    folder_names = []
    for folder in os.listdir(CWD):
        if os.path.isdir(os.path.join(CWD, folder)):
            folder_names.append(os.path.join(CWD, folder))
    
    excel = EXCEL(os.path.join(CWD, "通知信息.xlsx"), False)
    excel.active_ws_by_name(excel.get_sheets_name[0])

    # assertion
    # assert len(folder_names) == excel.rows - 1

    for folder in tqdm.tqdm(folder_names):
        name = folder.split(os.path.sep)[-1]
        save_name = get_pos_name_pair(excel, name)
        file_list = []
        bookmarks_list = []
        for f in os.listdir(os.path.join(CWD, name)):
            if os.path.isfile(os.path.join(folder, f)):
                file_list.append(os.path.join(folder, f))
                bookmarks_list.append(f.split(os.path.extsep)[0])

        """
        每个人要整合的报告是11份，整合顺序为：
            1.综合简版报告
            2.锐途管理人员人事决策报告
            3.心理风险计算机自适应测评面试报告
            4.团队角色评估报告
            5.管理个性V2测评报告
            6.管理风格测评报告
            7.FAST高潜人才评估报告
            8.管理人员偏离因素测评报告
            9.职业锚测评报告
            10.锐途管理人员自我发展报告
            11.锐途管理人员人事决策测评报告（简版）
        """
        pdf_merged = PyPDF2.PdfFileMerger()

        page_count = 0

        def merge_file_attach_tags(pdf_merged: PyPDF2.PdfFileMerger, page_count: int,  file_list, bookmarks_list, order: str) -> int:
            gen_ordered = "{}-{}".format(name, order)
            pdf_merged.add_bookmark(gen_ordered, page_count)
            orderd_file_path = file_list[bookmarks_list.index(gen_ordered)]
            pdf_reader = PyPDF2.PdfFileReader(stream=orderd_file_path)
            page_count += pdf_reader.getNumPages()
            pdf_merged.append(pdf_reader)
            return page_count

        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "综合简版报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "锐途管理人员人事决策报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "心理风险计算机自适应测评面试报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "团队角色评估报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "管理个性V2测评报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "管理风格测评报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "FAST高潜人才评估报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "管理人员偏离因素测评报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "职业锚测评报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "锐途管理人员自我发展报告")
        page_count = merge_file_attach_tags(pdf_merged, page_count, file_list, bookmarks_list, "锐途管理人员人事决策测评报告（简版）")
        
        # save to pdf format
        pdf_merged.write(save_name + ".pdf")
    
    print("全部合并完成，转换数量：{}".format(len(folder_names)))


if __name__ == '__main__':
    main()
